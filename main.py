import csv
import re

import bonobo
import openpyxl
import requests


def extract_excel():
    URL = "http://www.mizaniatouna.gov.tn/tunisia/template_fr/boost_fr.xlsx"
    FILE_NAME = "boost_db.xlsx"
    r = requests.get(URL)
    with open(FILE_NAME, "wb") as excel_f:
        excel_f.write(r.content)
    work_book = openpyxl.load_workbook(FILE_NAME, read_only=True)
    work_sheet = work_book.get_sheet_by_name("Sheet1")
    yield work_sheet


def transform_rows(work_sheet):
    header_regex = re.compile("[A-Z]+\d*\s*")
    # header
    raw_header_row = work_sheet.iter_rows(max_row=1)
    # TODO: better regex, this is hacky
    header_row = [header_regex.search(cell.value).group(
    ).strip() for cell in next(raw_header_row)]
    # rows
    rows = work_sheet.iter_rows(min_row=2, max_row=3)
    while True:
        yield dict(zip(header_row, [cell.value for cell in next(rows)]))


graph = bonobo.Graph(
    extract_excel,
    transform_rows,
    # TODO: replace separator with comma
    bonobo.CsvWriter("out.csv", ioformat="arg0")
)


if __name__ == '__main__':
    from bonobo.commands.run import get_default_services
    bonobo.run(graph, services=get_default_services(__file__))
